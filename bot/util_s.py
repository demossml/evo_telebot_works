from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import openpyxl
import traceback
import io
from numbers_parser import Document

from numbers_parser import Document
import traceback
import logging
from arrow import utcnow, get
from bd.model import Shop, AfsRequest, Employees, GetTime, Status


import logging

logger = logging.getLogger(__name__)


def format_message_list2(obj):
    text = ""  # Создаем пустую строку, в которую будем добавлять текст
    messages = []  # Создаем пустой список для хранения сообщений

    if len(obj) > 0:  # Проверяем, что входной объект не пуст
        for k, v in obj.items():  # Проходим по ключам и значениям в объекте
            key = str(k)  # Преобразуем ключ в строку
            val = str(v)  # Преобразуем значение в строку
            total_len = len(key) + len(val)  # Вычисляем общую длину ключа и значения
            pad = 31 - total_len % 31  # Вычисляем количество пробелов для выравнивания

            text += key  # Добавляем ключ в текст

            if pad > 0:
                text += " " * pad  # Добавляем пробелы для выравнивания

            if total_len > 31:
                text += " " * 2  # Добавляем двойные пробелы, если общая длина больше 31

            text += str(v)  # Добавляем значение в текст
            text += "\n"  # Добавляем символ новой строки

        # Разбиваем текст на части, если он слишком большой
        index = 0
        size = 4000
        while len(text) > 0:
            part = text[index : index + size]  # Вырезаем часть текста заданного размера
            index = part.rfind("\n")  # Находим последний символ новой строки в части
            if index == -1:
                index = len(
                    text
                )  # Если символ новой строки не найден, используем конец текста
            part = text[0:index]  # Выбираем часть текста до символа новой строки
            messages.append(
                "```\n" + part + "\n```"
            )  # Добавляем часть текста в список сообщений
            text = text[
                index:
            ].strip()  # Удаляем обработанную часть из текста и убираем пробелы

    return messages  # Возвращаем список сообщений


def format_message_list4(obj):
    text = ""  # Создаем пустую строку для хранения текста сообщений.
    messages = []  # Создаем пустой список для хранения отформатированных сообщений.

    if len(obj) > 0:  # Проверяем, есть ли объекты в списке.
        for i in obj:  # Проходим по каждому объекту в списке.
            for k, v in i.items():  # Проходим по каждой паре ключ-значение в объекте.
                key = str(k)  # Преобразуем ключ в строку.
                val = str(v)  # Преобразуем значение в строку.
                total_len = len(key) + len(
                    val
                )  # Вычисляем общую длину ключа и значения.
                pad = (
                    30 - total_len % 30
                )  # Вычисляем количество пробелов, чтобы выровнять текст.

                text += key  # Добавляем ключ к тексту.

                if pad > 0:  # Если нужно добавить пробелы для выравнивания,
                    text += " " * pad  # добавляем их.

                if total_len > 30:  # Если общая длина превышает 30 символов,
                    text += " " * 2  # добавляем 2 дополнительных пробела.

                text += str(v)  # Добавляем значение к тексту.
                text += "\n"  # Добавляем перевод строки между ключами и значениями.
            text += "\n"  # Добавляем пустую строку после каждого объекта.
            text += "******************************"  # Добавляем разделительную строку.
            text += "\n"

        text += ""  # Пустая строка (это выглядит как ошибка, потому что она ничего не делает).
        index = 0  # Начальный индекс для разделения текста на части.
        size = 4000  # Максимальная длина каждой части сообщения.
        while len(text) > 0:  # Пока есть текст для обработки:
            part = text[
                index : index + size
            ]  # Выбираем часть текста длиной не более 4000 символов.
            index = part.rfind(
                "\n"
            )  # Находим последний символ перевода строки в части.
            if index == -1:  # Если символ перевода строки не найден,
                index = len(text)  # используем всю часть текста.
            part = text[
                0:index
            ]  # Выбираем часть текста до найденного символа перевода строки.
            messages.append(
                "```\n" + part + "\n```"
            )  # Добавляем часть текста в список сообщений,
            text = text[index:].strip()  # и удаляем ее из исходного текста.

        return messages  # Возвращаем список отформатированных сообщений.


# Инициализация логгера
logger = logging.getLogger(__name__)


def xls_to_json_format_change(downloaded_file, file_format):
    try:
        logger.info("Начало преобразования файла в JSON")

        if file_format == ".xls" or file_format == ".xlsx":
            # Создаем буферизированный объект для чтения из загруженного файла
            file_buffer = io.BytesIO(downloaded_file)

            # Открываем книгу Excel
            book = openpyxl.load_workbook(file_buffer)

            # Получаем активный лист из книги Excel
            ws = book.active

            my_list = []  # Создаем пустой список для хранения словарей

            # Находим номер последнего столбца и строки
            last_column = len(list(ws.columns))
            last_row = len(list(ws.rows))

            # Проходимся по каждой строке в таблице Excel
            for row in range(1, last_row + 1):
                my_dict = {}  # Создаем пустой словарь для текущей строки
                # Проходимся по каждому столбцу в текущей строке
                for column in range(1, last_column + 1):
                    column_letter = get_column_letter(
                        column
                    )  # Получаем буквенное обозначение столбца

                    if row > 1:  # Пропускаем первую строку, так как это заголовки
                        # Добавляем элементы в словарь в формате "значение заголовка: значение ячейки"
                        header = str(
                            ws[column_letter + str(1)].value
                        ).strip()  # Убираем пробелы в заголовках
                        cell_value = ws[column_letter + str(row)].value
                        if isinstance(cell_value, str):
                            cell_value = (
                                cell_value.strip()
                            )  # Убираем пробелы в значениях
                        my_dict[header] = cell_value
                if len(my_dict) > 0:  # Убеждаемся, что словарь не пустой
                    my_list.append(my_dict)  # Добавляем словарь в список

        elif file_format == ".numbers":
            # Создаем буферизированный объект для чтения из загруженного файла
            file_buffer = io.BytesIO(downloaded_file)

            # Открываем документ .numbers
            doc = Document(file_buffer)
            sheets = doc.sheets()
            table = sheets[0].tables()[0]  # Используем первую таблицу в первом листе

            my_list = []  # Создаем пустой список для хранения словарей

            # Получаем данные из таблицы
            rows = table.rows()
            headers = rows[0]  # Первая строка - заголовки

            # Проходимся по каждой строке в таблице .numbers
            for row in rows[1:]:  # Пропускаем первую строку с заголовками
                my_dict = {}  # Создаем пустой словарь для текущей строки
                for header, cell in zip(headers, row):
                    header = header.strip()  # Убираем пробелы в заголовках
                    cell_value = cell.value
                    if isinstance(cell_value, str):
                        cell_value = cell_value.strip()  # Убираем пробелы в значениях
                    my_dict[header] = cell_value
                if len(my_dict) > 0:  # Убеждаемся, что словарь не пустой
                    my_list.append(my_dict)  # Добавляем словарь в список

        logger.info("Преобразование завершено")
        return my_list  # Возвращаем список словарей

    except openpyxl.utils.exceptions.InvalidFileException:
        logger.error("Загруженный файл не является действительным файлом Excel.")
        return None
    except Exception as e:
        logger.error(f"Произошла ошибка: {e}")
        traceback.print_exc()
        return None


def status_shop(shop_id: str) -> bool:
    # Получаем объект статуса из базы данных для указанного магазина со статусом "deleted"
    doc_status = Status.objects(shop=shop_id, status="deleted").first()
    # Возвращаем True, если объект не найден или его статус "restore", иначе возвращаем False
    return not (doc_status and doc_status.status != "restore")


# # Функция для отправки сообщений по расписанию
def send_scheduled_message():
    since = utcnow().replace(hour=2).isoformat()

    result = {}
    for shop in Shop.objects():
        if status_shop(shop["uuid"]):
            documents = GetTime.objects(
                __raw__={"openingData": {"$gte": since}, "shopUuid": shop["uuid"]}
            )

            result[shop["name"]] = "ЕЩЕ НЕ ОТКРЫТА!!!"

            for doc in documents:
                user_id = str(doc.user_id)
                employees = [
                    element["name"]
                    for element in Employees.objects(lastName=user_id).only("name")
                ]
                if doc["openingData"]:
                    result[shop["name"]] = "{} {}".format(
                        employees[0], doc["openingData"][11:16]
                    )
    return [result]
